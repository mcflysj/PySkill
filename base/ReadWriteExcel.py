"""
Excel文件读取与写入

Version: 0.1
Author: Mcfly_SJ
Date: 2020-12-26
"""
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook

# start write excel
workbook = Workbook()
sheet = workbook.active
data = [
    [1001, '鸣人', '木叶村', '螺旋丸'],
    [1002, '佐助', '木叶村', '千鸟']
]
sheet.append(['编号', '姓名', '所属村', '绝招'])
for row in data:
    sheet.append(row)
tab = Table(displayName="Table1", ref="A1:E5")

tab.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium7", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=True)
sheet.add_table(tab)
workbook.save('../data/excelDemo.xlsx')
# end write excel

# start read excel
workbook = load_workbook('../data/excelDemo.xlsx')
print(workbook.sheetnames)
sheet = workbook[workbook.sheetnames[0]]
print(sheet.title)
for row in range(2, 7):
    for col in range(65, 70):
        cell_index = chr(col) + str(row)
        print(sheet[cell_index].value, end='\t')
    print()
# end read excel

